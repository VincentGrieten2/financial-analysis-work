import pdfplumber
import re
import os
import csv
from pathlib import Path
from decimal import Decimal
from collections import defaultdict
from openpyxl import load_workbook
from datetime import datetime, timedelta
import logging
import shutil
import openpyxl
from openpyxl.utils import get_column_letter
import sys

# Configureer logging
handlers = [logging.FileHandler('parser.log')]
# Only add StreamHandler if not running as executable
if not getattr(sys, 'frozen', False):
    handlers.append(logging.StreamHandler())

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=handlers
)

# ==============================================
# CATEGORIE DEFINITIES
# ==============================================

# RESULTATENREKENING CATEGORIEËN
resultaten_cat = [
    # Bedrijfsopbrengsten
    ("A", "70", "Omzet"),
    ("B", "71", "Voorraadwijzigingen (+) (-)"),
    ("C", "72", "Geproduceerde vaste activa"),
    ("D", "74", "Andere bedrijfsopbrengsten"),
    
    # Bedrijfskosten
    ("E", "600/8", "Handelsg., grond & hulpst."),
    ("F", "609", "Wijzigingen in de voorraad"),
    ("G", "61", "Diensten en diverse goed."),
    ("H", "62", "Bezoldig., soc. lasten en pensioen."),
    ("I", "630", "Afschrijvingen en waardevermi."),
    ("J", "631/4", "Waardevermind. op voorraden, b.i.u."),
    ("K", "635/7", "Voorz.vr risico's en kosten"),
    ("L", "640/8", "Andere bedrijfskosten"),
    ("M", "649", "Geactiveerde bedrijfskosten (-)"),
    
    # Financiële opbrengsten
    ("N", "750", "Opbrengsten uit fin. vaste activa"),
    ("O", "751", "Opbrengsten uit vlottende activa"),
    ("P", "752/9", "Andere financiële opbrengsten"),
    
    # Financiële kosten
    ("Q", "650", "Kosten van schulden"),
    ("R", "651", "Waardeverminderingen op vlottende act."),
    ("S", "652/9", "Andere financiële kosten"),
    
    # Uitzonderlijke opbrengsten
    ("T", "760", "Terugnem. afschr.en waardevermind. (im)mat.v.a."),
    ("U", "761", "Terugneming waardeverm.op fin.v.a."),
    ("V", "762", "Terugnem. voorz.vr uitz. kosten"),
    ("W", "763", "Meerwaarde bij realis. vaste act."),
    ("X", "764/9", "And. uitzonderlijke opbrengsten"),
    
    # Uitzonderlijke kosten
    ("Y", "660", "Uitz. afschrijvingen"),
    ("Z", "661", "Waardeverminder. op fin. vaste act."),
    ("AA", "662", "Voorz. vr uitz. risico's en kosten"),
    ("AB", "663", "Minderw. bij realisatie vaste act."),
    ("AC", "664/8", "Andere uitzonderlijke kosten"),
    ("AD", "669", "Geactiveerde uitzonderlijke kosten (-)"),
    
    # Belastingen
    ("AE", "670/3", "Belastingen (-)"),
    ("AF", "77", "Regulariser. vn belast. & terugnemingen vn voorziening. voor belast."),
]

# BALANS CATEGORIEËN
balans_cat = [
    # VASTE ACTIVA
    ("AZ", "20", "Oprichtingskosten"),
    ("BA", "21", "Immateriële vaste activa"),
    ("BB", "22", "Terreinen en gebouwen"),
    ("BC", "23", "Installaties, machines en uitrust."),
    ("BD", "24", "Meubilair en rollend mat."),
    ("BE", "25", "Leasing en soortgelijke rechten"),
    ("BF", "26", "Overige materiele vaste activa"),
    ("BG", "27", "Activa in aanbouw en vooruitbetal."),
    
    # FINANCIËLE VASTE ACTIVA
    ("BH", "280", "Verbonden ondernemingen - Deelnemingen"),
    ("BI", "281", "Verbonden ondernemingen - Vorderingen"),
    ("BJ", "282", "Ondernemingen met deelnemingsverh. - Deelnemingen"),
    ("BK", "283", "Ondernemingen met deelnemingsverh. - Vorderingen"),
    ("BL", "284", "Andere financiele vaste activa - Aandelen"),
    ("BM", "285/8", "Andere financiele vaste activa - Vorder. & borgtochten"),
    
    # VLOTTENDE ACTIVA
    ("BN", "290", "Handelsvorderingen"),
    ("BO", "291", "Overige vorderingen"),
    
    # VOORRADEN
    ("BP", "30/31", "Grond- en hulpstoffen"),
    ("BQ", "32", "Goederen in bewerking"),
    ("BR", "33", "Gereed produkt"),
    ("BS", "34", "Handelsgoederen"),
    ("BT", "35", "Onroer. goed. bestemd vr. verkoop"),
    ("BU", "36", "Vooruitbetalingen"),
    ("BV", "37", "Bestellingen in uitvoering"),
    
    # OVERIGE VLOTTENDE ACTIVA
    ("BW", "40", "Handelsvorderingen"),
    ("BX", "41", "Overige vorderingen"),
    ("BY", "50", "Eigen aandelen"),
    ("BZ", "51/53", "Overige beleggingen"),
    ("CA", "54/58", "Liquide middelen"),
    ("CB", "490/1", "Overlopende rekeningen"),
    
    # EIGEN VERMOGEN
    ("CC", "100", "Geplaatst kapitaal"),
    ("CD", "101", "Niet opgevraagd kapitaal (-)"),
    ("CE", "11", "Uitgiftepremies"),
    ("CF", "12", "Herwaarderingsmeerwaarden"),
    
    # RESERVES
    ("CG", "130", "Wettelijke reserve"),
    ("CH", "131", "Onbeschikbare reserves"),
    ("CI", "132", "Belastingvrije reserves"),
    ("CJ", "133", "Beschikbare reserves"),
    ("CJ2", "14", "Overgedragen resultaat"),
    ("CK", "140", "Overgedragen winst"),
    ("CL", "141", "Overgedragen verlies (-)"),
    ("CM", "15", "Kapitaalsubsidies"),
    
    # VOORZIENINGEN EN UITGESTELDE BELASTINGEN
    ("CN", "160", "Pensioen. & soortgel. verplichting"),
    ("CO", "161", "Belastingen"),
    ("CP", "162", "Grote herstel- en onderhoudswerken"),
    ("CQ", "163/9", "Overige risico's & kosten"),
    ("CR", "168", "Uitgestelde belastingen"),
    
    # VREEMD VERMOGEN LANG
    ("CS", "170", "Achtergestelde leningen"),
    ("CT", "171", "Niet achtergestelde obl. leningen"),
    ("CU", "172", "Leasingschulden"),
    ("CV", "173", "Kredietinstellingen"),
    ("CW", "174", "Overige leningen"),
    
    # HANDELSSCHULDEN
    ("CX", "1750", "Handelsschulden - Leveranciers"),
    ("CY", "1751", "Handelsschulden - Te betalen wissels"),
    ("CZ", "176", "Ontvangen vooruitbetal. op bestel."),
    ("DA", "178/9", "Overige schulden"),
    
    # VREEMD VERMOGEN KORT
    ("DB", "42", "Schuld + 1 jr,vervallen - 1 jr"),
    ("DC", "430/8", "Financiele schulden - Kredietinstellingen"),
    ("DD", "439", "Financiele schulden - Overige leningen"),
    ("DE", "440/4", "Handelsschulden - Leveranciers"),
    ("DF", "441", "Handelsschulden - Te betalen wissels"),
    ("DG", "46", "Ontvangen vooruitbetal. op bestell."),
    ("DH", "450/3", "Belastingen"),
    ("DI", "454/9", "Bezoldigingen & sociale lasten"),
    ("DJ", "47/48", "Overige schulden"),
    ("DK", "492/3", "Overlopende rekeningen")
]

# ==============================================
# HELPER FUNCTIES
# ==============================================

def maak_categorie_mapping(categorieen):
    """Maak een mapping van rekeningcodes naar categorieën"""
    try:
        mapping = defaultdict(list)
        for cat_letter, rek_codes, omschrijving in categorieen:
            if '/' in rek_codes:  # Voor bereiken zoals 631/4
                start, end = rek_codes.split('/')
                if len(start) != len(end):
                    end = start[:-len(end)] + end  # Voor 752/9 → 752-759
                base = start[:-(len(end)-len(start)+1)] if len(start) > len(end) else start[:-len(end)]
                for i in range(int(start[-len(end):]), int(end)+1):
                    full_code = f"{base}{i:0{len(end)}d}"
                    mapping[full_code].append((cat_letter, omschrijving))
            else:
                mapping[rek_codes].append((cat_letter, omschrijving))
        return mapping
    except Exception as e:
        logging.error(f"Fout bij maken categorie mapping: {str(e)}")
        raise

def parse_bedrag(waarde_str):
    """Parse een bedrag string naar een Decimal object"""
    try:
        if not waarde_str:
            return Decimal("0")
        
        # Verwijder alle spaties
        waarde_str = waarde_str.strip()
        
        # Verwerk negatieve bedragen tussen haakjes
        is_negatief = False
        if waarde_str.startswith('(') and waarde_str.endswith(')'):
            waarde_str = waarde_str[1:-1]
            is_negatief = True
        
        # Als er een punt is en daarna een komma, dan is het punt een duizendtalscheiding
        if '.' in waarde_str and ',' in waarde_str:
            waarde_str = waarde_str.replace('.', '')
        
        # Vervang komma door punt voor decimaalteken
        waarde_str = waarde_str.replace(',', '.')
        
        waarde = Decimal(waarde_str)
        if is_negatief:
            waarde = -waarde
        return waarde
    except Exception as e:
        logging.error(f"Fout bij parsen van bedrag '{waarde_str}': {str(e)}")
        return Decimal("0")

def vind_categorie(rekening, mapping):
    """Vind de categorie voor een gegeven rekeningcode"""
    try:
        # Remove any decimal points from the account number
        base_rekening = rekening.split('.')[0]
        
        # Try exact matches first (e.g., 630), then more general ones (e.g., 63)
        # Sort by code length in descending order to try longer matches first
        for code_len in sorted({len(code) for code in mapping.keys()}, reverse=True):
            # Take only the first digits up to code_len
            code = base_rekening[:code_len]
            if code in mapping:
                return mapping[code][0]  # return (letter, omschrijving)
        return None, None
    except Exception as e:
        logging.error(f"Fout bij vinden categorie voor rekening {rekening}: {str(e)}")
        return None, None

def vind_datum_in_pdf(pdf_pad):
    """Zoek naar een datum in het PDF bestand"""
    try:
        # Day-level date formats
        dag_formaten = [
            ('%d-%m-%Y', r'\b\d{2}-\d{2}-\d{4}\b'),  # DD-MM-YYYY
            ('%d/%m/%Y', r'\b\d{2}/\d{2}/\d{4}\b'),  # DD/MM/YYYY
            ('%Y-%m-%d', r'\b\d{4}-\d{2}-\d{2}\b'),  # YYYY-MM-DD
            ('%Y/%m/%d', r'\b\d{4}/\d{2}/\d{2}\b'),  # YYYY/MM/DD
            ('%d.%m.%Y', r'\b\d{2}\.\d{2}\.\d{4}\b'),  # DD.MM.YYYY
            ('%d %B %Y', r'\b\d{1,2} [A-Za-z]+ \d{4}\b'),  # DD Month YYYY
        ]
        
        # Month-level date formats
        maand_formaten = [
            ('%m-%Y', r'\b\d{2}-\d{4}\b'),  # MM-YYYY
            ('%m/%Y', r'\b\d{2}/\d{4}\b'),  # MM/YYYY
            ('%Y-%m', r'\b\d{4}-\d{2}\b'),  # YYYY-MM
            ('%Y/%m', r'\b\d{4}/\d{2}\b'),  # YYYY/MM
            ('%m.%Y', r'\b\d{2}\.\d{4}\b'),  # MM.YYYY - Fixed escape sequence
        ]

        # Additional patterns to check for Belgian date formats
        belgian_date_patterns = [
            r'\b(\d{1,2})(?:st|nd|rd|th)? ([A-Za-z]+) (\d{4})\b',  # 31st December 2023
            r'\bper (\d{1,2})(?:\.| )(\d{1,2})(?:\.| )(\d{4})\b',  # per 31.12.2023 or per 31 12 2023
            r'\bafgesloten op (\d{1,2})(?:\.| )(\d{1,2})(?:\.| )(\d{4})\b',  # afgesloten op 31.12.2023
        ]
        
        gevonden_datums = []
        
        # Common Belgian months in Dutch and French for date parsing
        months_nl = {
            'januari': 1, 'februari': 2, 'maart': 3, 'april': 4, 'mei': 5, 'juni': 6,
            'juli': 7, 'augustus': 8, 'september': 9, 'oktober': 10, 'november': 11, 'december': 12
        }
        
        months_fr = {
            'janvier': 1, 'février': 2, 'mars': 3, 'avril': 4, 'mai': 5, 'juin': 6, 
            'juillet': 7, 'août': 8, 'septembre': 9, 'octobre': 10, 'novembre': 11, 'décembre': 12
        }

        with pdfplumber.open(pdf_pad) as pdf:
            all_text = []
            for page in pdf.pages[:3]:  # Only check first 3 pages to find date faster
                try:
                    text = page.extract_text()
                    if text:
                        all_text.append(text)
                except Exception as e:
                    logging.warning(f"Waarschuwing bij zoeken datum: {str(e)}")
                    continue

            text = "\n".join(all_text)
            
            # First try day-level formats
            for fmt, pattern in dag_formaten:
                matches = re.findall(pattern, text)
                for match in matches:
                    try:
                        datum = datetime.strptime(match, fmt)
                        if 1900 <= datum.year <= 2100:  # Reasonable year range
                            # Check if it's the last day of the month
                            volgende_maand = datum.replace(day=28) + timedelta(days=4)
                            laatste_dag = volgende_maand - timedelta(days=volgende_maand.day)
                            if datum.day == laatste_dag.day:
                                gevonden_datums.append(datum)
                    except ValueError:
                        continue

            # Check for Belgian specific date patterns
            for pattern in belgian_date_patterns:
                matches = re.findall(pattern, text, re.IGNORECASE)
                for match in matches:
                    try:
                        if len(match) == 3:  # Day, Month, Year
                            day = int(match[0])
                            # Handle both numeric month and text month
                            if match[1].isdigit():
                                month = int(match[1])
                            else:
                                month_text = match[1].lower()
                                if month_text in months_nl:
                                    month = months_nl[month_text]
                                elif month_text in months_fr:
                                    month = months_fr[month_text]
                                else:
                                    continue
                            year = int(match[2])
                            
                            if 1 <= day <= 31 and 1 <= month <= 12 and 1900 <= year <= 2100:
                                datum = datetime(year, month, day)
                                # Check if it's the last day of the month
                                volgende_maand = datum.replace(day=28) + timedelta(days=4)
                                laatste_dag = volgende_maand - timedelta(days=volgende_maand.day)
                                if datum.day == laatste_dag.day:
                                    gevonden_datums.append(datum)
                    except Exception:
                        continue

            # If no day-level dates found, try month-level formats
            if not gevonden_datums:
                for fmt, pattern in maand_formaten:
                    matches = re.findall(pattern, text)
                    for match in matches:
                        try:
                            # Parse the month and year
                            datum = datetime.strptime(match, fmt)
                            # Convert to last day of the month
                            volgende_maand = datum.replace(day=28) + timedelta(days=4)
                            laatste_dag = volgende_maand - timedelta(days=volgende_maand.day)
                            gevonden_datums.append(laatste_dag)
                        except ValueError:
                            continue

            # If still no dates found, check for year patterns
            if not gevonden_datums:
                current_year = datetime.now().year
                year_patterns = [
                    r'\b(20\d{2})\b',  # Find years in 2000-2099 range
                    r'\bBoekjaar (\d{4})\b',  # Boekjaar 2023
                    r'\bJaarrekening (\d{4})\b',  # Jaarrekening 2023
                ]
                
                for pattern in year_patterns:
                    matches = re.findall(pattern, text, re.IGNORECASE)
                    for match in matches:
                        try:
                            year = int(match)
                            if 1900 <= year <= current_year + 1:  # Reasonable year range
                                # Assume it's the end of the year (31st December)
                                datum = datetime(year, 12, 31)
                                gevonden_datums.append(datum)
                        except Exception:
                            continue

            if gevonden_datums:
                # Return the most recent date found
                return max(gevonden_datums)
                
            # No date found
            logging.warning(f"Geen datum gevonden in {pdf_pad}")
            return None

    except Exception as e:
        logging.error(f"Fout bij zoeken datum: {str(e)}")
        return None

def verwerk_pdf_sectie(pdf_pad, mapping, categorieen):
    """Verwerk een PDF sectie en extraheer de bedragen per categorie"""
    try:
        sommen = {cat[0]: Decimal("0") for cat in categorieen}
        omschrijvingen = {cat[0]: cat[2] for cat in categorieen}
        rekening_codes = {cat[0]: cat[1] for cat in categorieen}
        
        # Extract account codes and check if this is NBB format
        rekeningcodes, is_nbb_format, parent_accounts = extract_rekeningcodes_from_pdf(pdf_pad)
        
        # Process each account code
        for rekening, (description, bedrag) in rekeningcodes.items():
            # Skip processing parent accounts in NBB format - we'll process leaf nodes only
            if is_nbb_format and rekening in parent_accounts:
                logging.debug(f"Skipping parent account {rekening} with value {bedrag}")
                continue
                
            # Find the category for this account
            cat_letter, _ = vind_categorie(rekening, mapping)
            if cat_letter:
                sommen[cat_letter] = bedrag
                logging.debug(f"Added {bedrag} to category {cat_letter} for account {rekening}")

        return sommen, omschrijvingen, rekening_codes
    except Exception as e:
        logging.error(f"Fout bij verwerken PDF sectie {pdf_pad}: {str(e)}")
        raise

def maak_dataframe(sommen, omschrijvingen, rekening_codes, prefix=""):
    """Maak een DataFrame van de verwerkte data"""
    try:
        data = []
        for cat_letter, omschrijving in omschrijvingen.items():
            totaal = sommen[cat_letter]
            
            if totaal < 0:
                bedrag_str = f"({abs(totaal):,.2f})".replace(".", ",").replace(",", ".", 1)
            else:
                bedrag_str = f"{totaal:,.2f}".replace(".", ",").replace(",", ".", 1)
            
            data.append({
                "Categorie": prefix + cat_letter,
                "Rekening": rekening_codes[cat_letter],
                "Omschrijving": omschrijving,
                "Totaal (€)": bedrag_str
            })
        return data
    except Exception as e:
        logging.error(f"Fout bij maken DataFrame: {str(e)}")
        raise

def convert_currency_to_float(currency_str):
    """Converteer Europese valuta string naar float"""
    try:
        # Verwijder alle spaties
        cleaned = currency_str.strip()
        
        # Verwerk negatieve bedragen tussen haakjes
        is_negatief = False
        if cleaned.startswith('(') and cleaned.endswith(')'):
            cleaned = cleaned[1:-1]
            is_negatief = True
            
        # Verwijder alle punten (duizendtalscheiding)
        cleaned = cleaned.replace('.', '')
        
        # Als er een komma is, gebruik deze als decimaalteken
        if ',' in cleaned:
            parts = cleaned.split(',')
            if len(parts) == 2:  # Normaal geval: één komma
                cleaned = parts[0] + '.' + parts[1]
            else:  # Meerdere komma's, gebruik de laatste als decimaalteken
                cleaned = ''.join(parts[:-1]) + '.' + parts[-1]
        else:
            # Als er geen komma is, veronderstel dat de laatste 2 cijfers centen zijn
            if len(cleaned) > 2:
                cleaned = cleaned[:-2] + '.' + cleaned[-2:]
                
        # Converteer naar float
        waarde = float(cleaned)
        
        # Pas het teken toe
        if is_negatief:
            waarde = -waarde
            
        return waarde
    except Exception as e:
        logging.error(f"Fout bij conversie van '{currency_str}': {str(e)}")
        raise

def extract_rekeningcodes_from_pdf(pdf_pad):
    """Extract unique rekeningcodes and their descriptions from a PDF file
    
    Returns:
        tuple: (rekeningcodes, is_nbb_format, parent_accounts)
               - rekeningcodes: Dict of account codes to (description, value) tuples
               - is_nbb_format: Boolean indicating if file was processed as NBB format
               - parent_accounts: Dict of parent account codes to list of child account codes
    """
    try:
        rekeningcodes = {}
        is_nbb_format = False
        
        with pdfplumber.open(pdf_pad) as pdf:
            all_text = []
            for page in pdf.pages:
                try:
                    text = page.extract_text()
                    if text:
                        # Split into lines and clean them
                        lines = [line.strip() for line in text.split('\n') if line.strip()]
                        all_text.extend(lines)
                except Exception as e:
                    logging.warning(f"Waarschuwing bij verwerken pagina: {str(e)}")
                    continue

            tekst = "\n".join(all_text)

        # Regular pattern (account code first, then description, then value)
        regular_pattern = r'^\s*(\d{2,12}(?:\.\d{1,3})?|\d{2,3}/\d{2,3})\s+([^0-9].*?)\s+(-?[\d.,]+(?:,\d{2}|\.\d{2})|\([\d.,]+(?:,\d{2}|\.\d{2})\))'
        
        # NBB pattern (description first, then account code, then value)
        nbb_pattern = r'^\s*([^0-9].*?)\s+(\d{2,12}(?:\.\d{1,3})?|\d{2,3}/\d{2,3})\s+(-?[\d.,]+(?:,\d{2}|\.\d{2})|\([\d.,]+(?:,\d{2}|\.\d{2})\))'
        
        # Keep track of unique codes to detect duplicates
        seen_codes = set()
        duplicate_found = False
        
        # First try with regular pattern
        found_matches = False
        
        for lijn in tekst.splitlines():
            # Skip empty lines
            if not lijn.strip():
                continue
                
            # Stop if we've found a duplicate section
            if duplicate_found:
                break
                
            match = re.match(regular_pattern, lijn.strip())
            if match:
                found_matches = True
                rekening = match.group(1)
                description = match.group(2).strip()
                bedrag_str = match.group(3)
                
                logging.debug(f"Found match in line (regular format): {lijn}")
                logging.debug(f"Rekening: {rekening}, Description: {description}, Bedrag: {bedrag_str}")
                
                if bedrag_str:
                    bedrag = parse_bedrag(bedrag_str)
                    
                    # Check for duplicates to detect appendix section
                    if rekening in seen_codes:
                        # If we see the same code with the same value, it's likely an appendix
                        if rekening in rekeningcodes and abs(rekeningcodes[rekening][1] - bedrag) < Decimal('0.01'):
                            logging.info(f"Found duplicate code with same value: {rekening}")
                            duplicate_found = True
                            break
                    
                    seen_codes.add(rekening)
                    
                    # Store or update the rekeningcode info
                    if rekening not in rekeningcodes or (description and not rekeningcodes[rekening][0]):
                        rekeningcodes[rekening] = (description, bedrag)
                        logging.debug(f"Stored/Updated code {rekening} with description: {description}")
        
        # If no matches found with regular pattern, try NBB pattern
        if not found_matches:
            logging.info(f"No matches found with regular pattern, trying NBB format for {pdf_pad}")
            seen_codes = set()  # Reset seen codes
            duplicate_found = False
            is_nbb_format = True
            
            for lijn in tekst.splitlines():
                # Skip empty lines
                if not lijn.strip():
                    continue
                    
                # Stop if we've found a duplicate section
                if duplicate_found:
                    break
                    
                match = re.match(nbb_pattern, lijn.strip())
                if match:
                    description = match.group(1).strip()
                    rekening = match.group(2)
                    bedrag_str = match.group(3)
                    
                    logging.debug(f"Found match in line (NBB format): {lijn}")
                    logging.debug(f"Description: {description}, Rekening: {rekening}, Bedrag: {bedrag_str}")
                    
                    if bedrag_str:
                        bedrag = parse_bedrag(bedrag_str)
                        
                        # Check for duplicates to detect appendix section
                        if rekening in seen_codes:
                            # If we see the same code with the same value, it's likely an appendix
                            if rekening in rekeningcodes and abs(rekeningcodes[rekening][1] - bedrag) < Decimal('0.01'):
                                logging.info(f"Found duplicate code with same value: {rekening}")
                                duplicate_found = True
                                break
                        
                        seen_codes.add(rekening)
                        
                        # Store or update the rekeningcode info
                        if rekening not in rekeningcodes or (description and not rekeningcodes[rekening][0]):
                            rekeningcodes[rekening] = (description, bedrag)
                            logging.debug(f"Stored/Updated code {rekening} with description: {description}")
        
        # Identify parent accounts if we have NBB format
        parent_accounts = {}
        if is_nbb_format:
            parent_accounts = identify_sum_accounts(rekeningcodes)

        logging.info(f"Extracted {len(rekeningcodes)} unique rekeningcodes from {pdf_pad}")
        return rekeningcodes, is_nbb_format, parent_accounts
    except Exception as e:
        logging.error(f"Fout bij extractie rekeningcodes uit PDF {pdf_pad}: {str(e)}")
        return {}, False, {}

def create_overview_sheet(wb, pdf_with_dates):
    """Create an overview sheet with account codes from all PDFs"""
    try:
        # Create a new sheet for the overview
        overview_sheet = wb.create_sheet(title='Overview')
        
        # Set up headers
        overview_sheet['A1'] = 'Account Code'
        overview_sheet['B1'] = 'Description'
        
        # Collect all unique account codes
        all_codes = set()
        
        # Process each file
        for pdf_path in pdf_with_dates:
            logging.info(f"Processing {pdf_path} for Overview sheet")
            try:
                if pdf_path.lower().endswith('.csv'):
                    # For CSV files, extract codes from the first column
                    with open(pdf_path, 'r', encoding='utf-8') as f:
                        reader = csv.reader(f)
                        for row in reader:
                            if len(row) >= 1:
                                account = row[0].strip('"')
                                if account.replace('/', '').isdigit():
                                    all_codes.add(account)
                else:
                    # For PDF files, use the existing extraction method
                    try:
                        codes = extract_rekeningcodes_from_pdf(pdf_path)
                        all_codes.update(codes)
                        logging.info(f"Found {len(codes)} codes in PDF {pdf_path}")
                    except Exception as e:
                        logging.error(f"Fout bij extractie rekeningcodes uit PDF {pdf_path}: {str(e)}")
            except Exception as e:
                logging.error(f"Error processing file {pdf_path} for Overview: {str(e)}")
                continue
        
        logging.info(f"Total unique codes to be added to Overview: {len(all_codes)}")
        
        # Sort the codes and add to sheet
        sorted_codes = sorted(all_codes)
        for idx, code in enumerate(sorted_codes, start=2):  # Start from row 2
            overview_sheet[f'A{idx}'] = code
        
        logging.info(f"Successfully created Overview sheet with {len(sorted_codes)} rows")
        return overview_sheet
        
    except Exception as e:
        logging.error(f"Error creating Overview sheet: {str(e)}")
        raise

def export_naar_template(resultaten_data, balans_data, template_pad, output_pad, kolom, datum=None, pdf_with_dates=None):
    """Exporteert alle categorieën naar template Excel met specifieke kolom, behoudt bestaande waarden en formules"""
    wb = None
    try:
        # Check if output file exists, if not copy template
        if not os.path.exists(output_pad):
            shutil.copy2(template_pad, output_pad)
            logging.info(f"Created new output file from template: {output_pad}")
        
        # Load the existing output file
        wb = openpyxl.load_workbook(
            filename=output_pad,
            data_only=False,
            read_only=False
        )
        
        if 'Historiek' not in wb.sheetnames:
            logging.error("Worksheet 'Historiek' niet gevonden in template")
            return False

        ws = wb['Historiek']
        
        # Vul de datum in als deze beschikbaar is
        if datum:
            datum_cel = f"{kolom}7"
            ws[datum_cel] = datum.strftime("%d-%m-%Y")
            logging.info(f"Datum {datum.strftime('%d-%m-%Y')} ingevuld in cel {datum_cel}")

        # Get rekening 60 value to determine sign conversion
        rekening_60_value = None
        for data in resultaten_data:
            if data['Rekening'] == '60':
                try:
                    rekening_60_value = convert_currency_to_float(data['Totaal (€)'])
                    break
                except:
                    pass

        # Verwerk alle data
        for data in resultaten_data + balans_data:
            cat = data['Categorie']
            rek = data['Rekening']
            waarde_str = data['Totaal (€)']
            
            # Account 14 is now mapped directly to 140 or 141 in parse_csv_file,
            # so we don't need to handle it specially here
            if rek == '14':
                continue
            
            # Bepaal de doelcel op basis van de rekeningcode
            if rek in categorie_naar_cel:
                base_cel = categorie_naar_cel[rek]
                doel_cel = base_cel.replace('C', kolom)
                
                # Controleer of de cel een formule bevat
                if ws[doel_cel].value is not None and isinstance(ws[doel_cel].value, str) and ws[doel_cel].value.startswith('='):
                    logging.warning(f"Cel {doel_cel} bevat een formule en wordt niet overschreven")
                    continue
                
                try:
                    # Convert the value
                    waarde = convert_currency_to_float(waarde_str)
                    
                    # Special case for '670/3' (Belastingen) - always make it negative
                    if rek == '670/3' and waarde > 0:
                        waarde = -waarde
                    
                    # Apply sign conversion logic for cost rows if rekening 60 is negative
                    row_num = int(doel_cel[1:]) if len(doel_cel) > 1 else 0
                    is_cost_row = (154 <= row_num <= 165) or (174 <= row_num <= 177) or (188 <= row_num <= 194)
                    
                    if rekening_60_value is not None and rekening_60_value < 0 and is_cost_row:
                        if waarde < 0:
                            waarde = abs(waarde)  # Make costs positive
                        else:
                            waarde = -waarde  # Make non-costs negative
                    
                    ws[doel_cel] = waarde
                    logging.info(f"Cel {doel_cel} bijgewerkt met waarde {waarde}")
                except Exception as e:
                    logging.error(f"Fout bij updaten cel {doel_cel}: {str(e)}")
                    continue

        # Create Overview sheet after processing all PDFs
        # Only create Overview sheet when processing the last PDF (most recent one, column F)
        if kolom == 'F' and pdf_with_dates:
            create_overview_sheet(wb, pdf_with_dates)

        # Save the workbook
        wb.save(output_pad)
        wb.close()
        logging.info(f"Workbook succesvol bijgewerkt: {output_pad}")
        return True
                
    except Exception as e:
        logging.error(f"Fout bij exporteren naar template: {str(e)}")
        return False

    finally:
        # Zorg ervoor dat alle workbooks gesloten zijn
        if wb:
            try:
                wb.close()
            except:
                pass

# Maak categorie mappings
try:
    resultaten_mapping = maak_categorie_mapping(resultaten_cat)
    balans_mapping = maak_categorie_mapping(balans_cat)
    logging.info("Categorie mappings succesvol aangemaakt")
except Exception as e:
    logging.error(f"Fout bij aanmaken categorie mappings: {str(e)}")
    raise

# Mapping van rekeningcodes naar Excel cellen
categorie_naar_cel = {
    # Resultatenrekening
    '70': 'C148',    # Omzet
    '71': 'C149',    # Voorraadwijzigingen
    '72': 'C150',    # Geproduceerde vaste activa
    '74': 'C151',    # Andere bedrijfsopbrengsten
    
    '600/8': 'C157',    # Handelsgoederen, grond- en hulpstoffen
    '609': 'C158',   # Wijzigingen in de voorraad
    '61': 'C159',    # Diensten en diverse goederen
    '62': 'C160',    # Bezoldigingen, sociale lasten en pensioenen
    '630': 'C161',   # Afschrijvingen en waardeverminderingen
    '631/4': 'C162', # Waardeverminderingen op voorraden
    '635/7': 'C163', # Voorzieningen voor risico's en kosten
    '640/8': 'C164', # Andere bedrijfskosten
    '649': 'C165',   # Geactiveerde bedrijfskosten
    
    '750': 'C170',   # Opbrengsten uit financiële vaste activa
    '751': 'C171',   # Opbrengsten uit vlottende activa
    '752/9': 'C172', # Andere financiële opbrengsten
    
    '650': 'C175',   # Kosten van schulden
    '651': 'C176',   # Waardeverminderingen op vlottende activa
    '652/9': 'C177', # Andere financiële kosten
    
    '760': 'C182',   # Terugneming van afschrijvingen
    '761': 'C183',   # Terugneming van waardeverminderingen op financiële vaste activa
    '762': 'C184',   # Terugneming van voorzieningen voor uitzonderlijke risico's
    '763': 'C185',   # Meerwaarden bij de realisatie van vaste activa
    '764/9': 'C186', # Andere uitzonderlijke opbrengsten
    
    '660': 'C189',   # Uitzonderlijke afschrijvingen
    '661': 'C190',   # Waardeverminderingen op financiële vaste activa
    '662': 'C191',   # Voorzieningen voor uitzonderlijke risico's
    '663': 'C192',   # Minderwaarden bij de realisatie van vaste activa
    '664/8': 'C193', # Andere uitzonderlijke kosten
    '669': 'C194',   # Geactiveerde uitzonderlijke kosten
    
    '670/3': 'C208', # Belastingen
    '77': 'C210',    # Regularisering van belastingen
    
    # Balans
    '20': 'C17',     # Oprichtingskosten
    '21': 'C19',     # Immateriële vaste activa
    '22': 'C22',     # Terreinen en gebouwen
    '23': 'C23',     # Installaties, machines en uitrusting
    '24': 'C24',     # Meubilair en rollend materieel
    '25': 'C25',     # Leasing en soortgelijke rechten
    '26': 'C26',     # Overige materiële vaste activa
    '27': 'C27',     # Activa in aanbouw en vooruitbetalingen
    
    '280': 'C31',    # Verbonden ondernemingen - Deelnemingen
    '281': 'C32',    # Verbonden ondernemingen - Vorderingen
    '282': 'C35',    # Ondernemingen met deelnemingsverhouding - Deelnemingen
    '283': 'C36',    # Ondernemingen met deelnemingsverhouding - Vorderingen
    '284': 'C38',    # Andere financiële vaste activa - Aandelen
    '285/8': 'C39',  # Andere financiële vaste activa - Vorderingen en borgtochten
    
    '290': 'C44',    # Handelsvorderingen
    '291': 'C45',    # Overige vorderingen
    
    '30/31': 'C49',  # Grond- en hulpstoffen
    '32': 'C50',     # Goederen in bewerking
    '33': 'C51',     # Gereed product
    '34': 'C52',     # Handelsgoederen
    '35': 'C53',     # Onroerende goederen bestemd voor verkoop
    '36': 'C54',     # Vooruitbetalingen
    '37': 'C55',     # Bestellingen in uitvoering
    
    '40': 'C58',     # Handelsvorderingen
    '41': 'C59',     # Overige vorderingen
    '50': 'C62',     # Eigen aandelen
    '51/53': 'C63',  # Overige beleggingen
    '54/58': 'C65',  # Liquide middelen
    '490/1': 'C67',  # Overlopende rekeningen
    
    '100': 'C78',    # Geplaatst kapitaal
    '11': 'C81',     # Uitgiftepremies
    '12': 'C83',     # Herwaarderingsmeerwaarden
    
    '130': 'C86',    # Wettelijke reserve
    '131': 'C87',    # Onbeschikbare reserves
    '132': 'C88',    # Belastingvrije reserves
    '133': 'C89',    # Beschikbare reserves
    '14': 'C90',     # Overgedragen resultaat
    '140': 'C91',    # Overgedragen winst
    '141': 'C92',    # Overgedragen verlies
    '15': 'C94',     # Kapitaalsubsidies
    
    '160': 'C99',    # Pensioenen en soortgelijke verplichtingen
    '161': 'C100',   # Belastingen
    '162': 'C101',   # Grote herstellings- en onderhoudswerken
    '163/9': 'C102', # Overige risico's en kosten
    '168': 'C103',   # Uitgestelde belastingen
    
    '170': 'C109',   # Achtergestelde leningen
    '171': 'C110',   # Niet-achtergestelde obligatieleningen
    '172': 'C111',   # Leasingschulden en soortgelijke
    '173': 'C112',   # Kredietinstellingen
    '174': 'C113',   # Overige leningen
    
    '1750': 'C115',  # Handelsschulden - Leveranciers
    '1751': 'C116',  # Handelsschulden - Te betalen wissels
    '176': 'C117',   # Ontvangen vooruitbetalingen op bestellingen
    '178/9': 'C118', # Overige schulden
    
    '42': 'C121',    # Schulden op meer dan één jaar die binnen het jaar vervallen
    '430/8': 'C123', # Kredietinstellingen
    '439': 'C124',   # Overige leningen
    '440/4': 'C126', # Leveranciers
    '441': 'C127',   # Te betalen wissels
    '46': 'C128',    # Ontvangen vooruitbetalingen op bestellingen
    '450/3': 'C130', # Belastingen
    '454/9': 'C131', # Bezoldigingen en sociale lasten
    '47/48': 'C132', # Overige schulden
    '492/3': 'C134'  # Overlopende rekeningen
}

def generate_unique_output_filename(base_name='financial_analysis'):
    """Generate a unique output filename by appending timestamp."""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return f'{base_name}_{timestamp}.xlsx'

def parse_csv_file(csv_path, mapping, categorieen):
    """Parse a CSV file and extract financial data"""
    try:
        sommen = defaultdict(Decimal)
        omschrijvingen = {cat[0]: cat[2] for cat in categorieen}  # Initialize with all categories
        rekening_codes = {cat[0]: cat[1] for cat in categorieen}  # Initialize with all codes
        
        # Dictionary to store all accounts and their values
        all_accounts = {}
        summary_accounts = {}
        
        # First, find the accounting period end date
        datum = None
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row in reader:
                if len(row) == 2 and row[0].strip('"') == "Accounting period end date":
                    try:
                        datum = datetime.strptime(row[1].strip('"'), "%Y-%m-%d")
                        break
                    except ValueError:
                        logging.warning(f"Could not parse date from CSV: {row[1]}")
        
        # First pass: collect all accounts and identify summary accounts
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row in reader:
                if len(row) != 2:  # Skip non-data rows
                    continue
                    
                account = row[0].strip('"')
                value_str = row[1].strip('"')
                
                # Skip if not a valid account number
                if not account.replace('/', '').isdigit():
                    continue
                    
                try:
                    value = Decimal(str(value_str).replace(',', '.'))
                except (ValueError, TypeError, decimal.InvalidOperation):
                    continue
                
                all_accounts[account] = value
                
                # Identify summary accounts (e.g., "600/8", "10/49")
                if '/' in account:
                    summary_accounts[account] = value
                    
                # Special handling for 600/8 - store it directly in the appropriate category
                if account == "600/8":
                    cat_letter, _ = vind_categorie(account, mapping)
                    if cat_letter:
                        sommen[cat_letter] = value
                        logging.debug(f"Added summary account {account} with value {value} to category {cat_letter}")
        
        # Handle special cases for accounts 76 and 66
        if "76" in all_accounts:
            # For account 76, map to category X (764/9)
            for cat in categorieen:
                if cat[1] == "764/9":
                    sommen[cat[0]] = all_accounts["76"]
                    logging.debug(f"Directly mapped account 76 with value {all_accounts['76']} to category {cat[0]} (764/9)")
                    break
        
        if "66" in all_accounts:
            # For account 66, map to category AC (664/8)
            for cat in categorieen:
                if cat[1] == "664/8":
                    sommen[cat[0]] = all_accounts["66"]
                    logging.debug(f"Directly mapped account 66 with value {all_accounts['66']} to category {cat[0]} (664/8)")
                    break
                    
        # Process standard accounts and their values to avoid losing data
        for account, value in all_accounts.items():
            cat_letter, _ = vind_categorie(account, mapping)
            if cat_letter:
                if account in ["22", "23", "24", "430/8"]:  # Preserve these accounts' values
                    sommen[cat_letter] = value
                    logging.debug(f"Preserved value for account {account} with value {value} to category {cat_letter}")
                
        # CORRECTIONS:
        
        # BH: "280" - Verbonden ondernemingen - Deelnemingen
        if "280" in all_accounts:
            for cat in categorieen:
                if cat[1] == "280":
                    sommen[cat[0]] = all_accounts["280"]
                    logging.debug(f"Mapped account 280 with value {all_accounts['280']} to category {cat[0]}")
                    break
                    
        # BM: "285/8" - Use this value and zero out others in range BI-BL
        if "285/8" in all_accounts:
            for cat in categorieen:
                if cat[1] == "285/8":
                    sommen[cat[0]] = all_accounts["285/8"]
                    logging.debug(f"Mapped account 285/8 with value {all_accounts['285/8']} to category {cat[0]}")
                    break
        
        # Zero out categories BI to BL (281-284)
        for cat in categorieen:
            if cat[1] in ["281", "282", "283", "284"]:
                sommen[cat[0]] = Decimal(0)
                logging.debug(f"Set category {cat[0]} ({cat[1]}) to 0")
                
        # BP to BU: Handle vooraden - use "34" for inventory
        if "34" in all_accounts:
            for cat in categorieen:
                # Map account 34 to BS category
                if cat[1] == "34":
                    sommen[cat[0]] = all_accounts["34"]
                    logging.debug(f"Mapped account 34 with value {all_accounts['34']} to category {cat[0]}")
                    break
                    
        # Zero out other inventory accounts
        for cat in categorieen:
            if cat[1] in ["30/31", "32", "33", "35", "36"] and cat[1] != "34":
                sommen[cat[0]] = Decimal(0)
                logging.debug(f"Set inventory category {cat[0]} ({cat[1]}) to 0")
        
        # CD-CF: "110" instead of individual accounts
        if "110" in all_accounts:
            for cat in categorieen:
                if cat[1] == "12":  # Map 110 value to herwaarderingsmeerwaarden (12)
                    sommen[cat[0]] = all_accounts["110"]
                    logging.debug(f"Mapped account 110 with value {all_accounts['110']} to category {cat[0]} (12)")
                    break
        
        # CG-CJ: Handle reserves - use "133" for available reserves
        if "133" in all_accounts:
            for cat in categorieen:
                if cat[1] == "133":
                    sommen[cat[0]] = all_accounts["133"]
                    logging.debug(f"Mapped account 133 with value {all_accounts['133']} to category {cat[0]}")
                    break
                    
        # Zero out other reserve accounts
        for cat in categorieen:
            if cat[1] in ["130", "131", "132"] and cat[1] != "133":
                sommen[cat[0]] = Decimal(0)
                logging.debug(f"Set reserve category {cat[0]} ({cat[1]}) to 0")
        
        # CJ2: "14" - Overgedragen resultaat - map to 140 if positive, 141 if negative
        if "14" in all_accounts:
            value = all_accounts["14"]
            if value >= 0:
                # If positive, map to 140 (Overgedragen winst)
                for cat in categorieen:
                    if cat[1] == "140":
                        sommen[cat[0]] = value
                        logging.debug(f"Mapped account 14 with positive value {value} to category {cat[0]} (140)")
                        break
            else:
                # If negative, map to 141 (Overgedragen verlies) as positive value
                for cat in categorieen:
                    if cat[1] == "141":
                        sommen[cat[0]] = abs(value)  # Store as positive in 141
                        logging.debug(f"Mapped account 14 with negative value {value} as positive to category {cat[0]} (141)")
                        break
        
        # CT to CW: "171" to "174" - handle "173" and zero out others
        if "173" in all_accounts:
            for cat in categorieen:
                if cat[1] == "173":
                    sommen[cat[0]] = all_accounts["173"]
                    logging.debug(f"Mapped account 173 with value {all_accounts['173']} to category {cat[0]}")
                    break
                    
        # Zero out other long-term loan accounts
        for cat in categorieen:
            if cat[1] in ["171", "172", "174"] and cat[1] != "173":
                sommen[cat[0]] = Decimal(0)
                logging.debug(f"Set loan category {cat[0]} ({cat[1]}) to 0")
        
        # Second pass: map accounts to categories
        for account, value in all_accounts.items():
            # Skip processing if this is a sub-account of a summary account that we'll use
            skip_account = False
            for summary_acc in summary_accounts:
                if '/' in summary_acc:
                    base, end = summary_acc.split('/')
                    if len(end) == 1:  # Case like "600/8"
                        start_num = int(base)
                        end_num = int(base[:-1] + end)
                        if account != summary_acc and account.isdigit():
                            acc_num = int(account)
                            if start_num <= acc_num <= end_num:
                                skip_account = True
                                break
            
            if skip_account:
                continue
            
            # Find the category for this account
            cat_letter, cat_desc = vind_categorie(account, mapping)
            if cat_letter:
                # Special handling for specific categories
                if account == "70":  # A: Revenue
                    sommen[cat_letter] = value
                elif account == "74":  # D: Other operating income (use main account)
                    sommen[cat_letter] = value
                elif account == "609":  # F: Direct match
                    sommen[cat_letter] = value
                elif account == "61":  # G: Direct match
                    sommen[cat_letter] = value
                elif account == "62":  # H: Use main account
                    sommen[cat_letter] = value
                elif account == "630":  # I: Direct match
                    sommen[cat_letter] = value
                elif account == "631/4":  # J: Use summary value
                    sommen[cat_letter] = value
                elif account == "640/8":  # L: Use summary value
                    sommen[cat_letter] = value
                elif account == "751":  # O: Direct match
                    sommen[cat_letter] = value
                elif account == "752/9":  # P: Use summary value
                    sommen[cat_letter] = value
                elif account == "650":  # Q: Direct match
                    sommen[cat_letter] = value
                elif account == "652/9":  # S: Use summary value
                    sommen[cat_letter] = value
                elif account == "670/3":  # AE: Use summary value
                    sommen[cat_letter] = value
                # Balance sheet categories that weren't handled in the special cases above
                elif account == "54/58":  # CA: Use summary value
                    sommen[cat_letter] = value
                elif account == "490/1":  # CB: Use summary value
                    sommen[cat_letter] = value
                elif account == "40":  # BW: Direct match
                    sommen[cat_letter] = value
                elif account == "41":  # BX: Direct match
                    sommen[cat_letter] = value
                elif account == "42":  # DB: Direct match
                    sommen[cat_letter] = value
                elif account == "46":  # DG: Direct match
                    sommen[cat_letter] = value
                elif account == "454/9":  # DI: Use summary value
                    sommen[cat_letter] = value
                elif account == "47/48":  # DJ: Use summary value
                    sommen[cat_letter] = value
                
                logging.debug(f"Added {value} to category {cat_letter} for account {account}")
        
        # Create the final dictionaries with all categories
        sommen_dict = {cat[0]: sommen.get(cat[0], Decimal("0")) for cat in categorieen}
        
        return sommen_dict, omschrijvingen, rekening_codes
    except Exception as e:
        logging.error(f"Error parsing CSV file {csv_path}: {str(e)}")
        raise

def process_file(file_path, mapping, categorieen):
    """Process either a PDF or CSV file and return the extracted data"""
    try:
        if file_path.lower().endswith('.pdf'):
            return verwerk_pdf_sectie(file_path, mapping, categorieen)
        elif file_path.lower().endswith('.csv'):
            return parse_csv_file(file_path, mapping, categorieen)
        else:
            raise ValueError(f"Unsupported file type: {file_path}")
    except Exception as e:
        logging.error(f"Error processing file {file_path}: {str(e)}")
        raise

def process_pdfs(template_path, input_files):
    """Process PDF and CSV files and export to template"""
    try:
        # Validate input files
        if not input_files:
            raise ValueError("No input files provided")
        if not template_path or not os.path.exists(template_path):
            raise ValueError("Invalid template path")
            
        # Create mappings
        resultaten_mapping = maak_categorie_mapping(resultaten_cat)
        balans_mapping = maak_categorie_mapping(balans_cat)
        
        # Process each file
        all_data = []
        pdf_with_dates = {}
        
        for file_path in input_files:
            try:
                if file_path.lower().endswith('.pdf'):
                    # Extract date from PDF
                    datum = vind_datum_in_pdf(file_path)
                    if datum:
                        pdf_with_dates[file_path] = datum
                elif file_path.lower().endswith('.csv'):
                    # For CSV files, extract date from the file content
                    with open(file_path, 'r', encoding='utf-8') as f:
                        reader = csv.reader(f)
                        for row in reader:
                            if len(row) == 2 and row[0].strip('"') == "Accounting period end date":
                                try:
                                    datum = datetime.strptime(row[1].strip('"'), "%Y-%m-%d")
                                    pdf_with_dates[file_path] = datum
                                    break
                                except ValueError:
                                    logging.warning(f"Could not parse date from CSV: {row[1]}")
                else:
                    logging.warning(f"Skipping unsupported file: {file_path}")
                    continue
                    
                # Process the file
                resultaten_data = process_file(file_path, resultaten_mapping, resultaten_cat)
                balans_data = process_file(file_path, balans_mapping, balans_cat)
                
                # Create data frames
                resultaten_df = maak_dataframe(*resultaten_data)
                balans_df = maak_dataframe(*balans_data, prefix="B")
                
                all_data.append((resultaten_df, balans_df))
                
            except Exception as e:
                logging.error(f"Error processing file {file_path}: {str(e)}")
                raise
            
        # Generate output filename
        output_path = generate_unique_output_filename()
        
        # Export data to template
        for idx, (resultaten_data, balans_data) in enumerate(all_data, start=1):
            kolom = get_column_letter(idx * 2)  # Skip a column between datasets
            export_naar_template(
                resultaten_data, balans_data,
                template_path, output_path,
                kolom,
                pdf_with_dates=pdf_with_dates
            )
            
        return output_path
    except Exception as e:
        logging.error(f"Error in process_pdfs: {str(e)}")
        raise

def identify_sum_accounts(rekeningcodes):
    """Identify parent/summary accounts that are the sum of other accounts.
    
    In the NBB format, many accounts are parent accounts that represent the sum
    of their child accounts. This function identifies these accounts by looking
    for accounts that have the same numeric pattern or are range-based (e.g., '22/27').
    
    Args:
        rekeningcodes: A dictionary where keys are account codes and values are tuples
                      of (description, value)
    
    Returns:
        A dictionary mapping parent account codes to a list of their child account codes.
    """
    try:
        # Sort account codes to help identify relationships
        sorted_codes = sorted(rekeningcodes.keys())
        
        # Map to store parent-child relationships
        parent_child_map = {}
        
        # First handle range-based account codes (e.g., '22/27', '54/58')
        for code in sorted_codes:
            if '/' in code:
                try:
                    start_code, end_code = code.split('/')
                    # Ensure both parts are valid numbers
                    if start_code.isdigit() and end_code.isdigit():
                        start_num = int(start_code)
                        end_num = int(end_code)
                        
                        # Find all child accounts in the range
                        children = []
                        for potential_child in sorted_codes:
                            # Skip the range code itself and non-numeric codes
                            if potential_child == code or '/' in potential_child:
                                continue
                                
                            # Get the base number (first 2 or 3 digits)
                            child_base = int(potential_child[:2] if len(potential_child) >= 2 else potential_child)
                            
                            # Check if the child's base number falls within the range
                            if start_num <= child_base <= end_num:
                                children.append(potential_child)
                        
                        if children:
                            parent_child_map[code] = children
                            logging.debug(f"Added range-based parent account {code} with {len(children)} children")
                except Exception as e:
                    logging.warning(f"Error processing range-based account {code}: {str(e)}")
                    continue
        
        # Check for base accounts (level 1 accounts like 20, 21, 22, etc.)
        base_accounts = {}
        for code in sorted_codes:
            # Skip range-based codes as they're already processed
            if '/' in code:
                continue
                
            # Check if this is a base account (2 digits)
            if len(code) == 2 and code.isdigit():
                base_accounts[code] = []
                
                # Find potential child accounts
                for child_code in sorted_codes:
                    if child_code != code and child_code.startswith(code) and '/' not in child_code:
                        base_accounts[code].append(child_code)
        
        # Check for level 2 accounts (3 digits, like 200, 201, etc.)
        level2_accounts = {}
        for code in sorted_codes:
            # Skip range-based codes as they're already processed
            if '/' in code:
                continue
                
            # Check if this is a level 2 account (3 digits)
            if len(code) == 3 and code.isdigit():
                level2_accounts[code] = []
                
                # Find potential child accounts
                for child_code in sorted_codes:
                    if child_code != code and child_code.startswith(code) and '/' not in child_code:
                        level2_accounts[code].append(child_code)
        
        # Merge maps
        parent_child_map.update(base_accounts)
        parent_child_map.update(level2_accounts)
        
        # Filter out parent accounts that don't have children
        parent_child_map = {parent: children for parent, children in parent_child_map.items() 
                           if children}
        
        # Validate parent-child relationships by checking sums
        validated_parents = {}
        for parent, children in parent_child_map.items():
            # Get the parent value
            if parent not in rekeningcodes:
                continue
                
            parent_value = rekeningcodes[parent][1]
            
            # Calculate the sum of child values
            child_sum = Decimal("0")
            all_children_present = True
            
            for child in children:
                if child in rekeningcodes:
                    child_value = rekeningcodes[child][1]
                    child_sum += child_value
                else:
                    all_children_present = False
                    break
            
            # Check if parent value equals sum of children (with small tolerance for rounding)
            if all_children_present and abs(parent_value - child_sum) < Decimal("0.1"):
                validated_parents[parent] = children
                logging.debug(f"Validated parent account {parent} with {len(children)} children")
            else:
                # Include if it's a range-based account (e.g., '22/27')
                if '/' in parent:
                    validated_parents[parent] = children
                    logging.debug(f"Included range-based parent account {parent} with {len(children)} children")
                # Also include it if it's clearly a parent account by structure (has 2-3 digits)
                elif len(parent) in (2, 3) and len(children) >= 2:
                    validated_parents[parent] = children
                    logging.debug(f"Included structural parent account {parent} with {len(children)} children")
        
        # Log identified parent accounts
        logging.info(f"Identified {len(validated_parents)} potential parent/summary accounts")
        
        return validated_parents
        
    except Exception as e:
        logging.error(f"Error identifying sum accounts: {str(e)}")
        return {}

if __name__ == "__main__":
    print("This module provides PDF financial statement processing functionality.")
    print("Please use run_financial_analysis.py to execute the processing.")